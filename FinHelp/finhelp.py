import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries
from openpyxl.cell.cell import MergedCell
from copy import copy

# ---------- Настройки файлов ----------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(CURRENT_DIR, "FinHelpExample.xlsx")

INSERT_START_ROW = 17
MAX_TABLE_COL = 8

KEYWORDS = ["финанс", "молияв"]

LEGAL_ENTITIES = {
    'OOO', 'MCHJ', 'MCHZ', 'YATT', 'XK', 'QK', 'AJ', 'PE', 'FE', 'IP',
    'MCH', 'LLC', 'LTD', 'YTT', 'CHP', 'SP', 'AO', 'DK', 'OK',
    'JAMIYATI', 'CHEKLANGAN', 'MASULIYATI', 'XUSUSIY', 'KORXONASI',
    'FAMILY', 'MARKAZ', 'UNIVERSAL', 'BIZNES', 'TRADE', 'SERVICE',
    'MCHJ', 'XK', 'YTT'
}


# ==========================================
# УТИЛИТЫ И ОЧИСТКА
# ==========================================
def sanitize_filename_part(s: str) -> str:
    if s is None: return "UNKNOWN"
    t = str(s).strip()
    t = re.sub(r'[\\/*?:"<>|]', "", t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t or "UNKNOWN"


def normalize_firm_name_key(name_str):
    if not name_str: return ""
    s = str(name_str).upper()
    translit_map = {
        'А': 'A', 'В': 'B', 'Е': 'E', 'К': 'K', 'М': 'M', 'Н': 'N', 'О': 'O',
        'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U', 'Х': 'H', 'Ё': 'YO', 'Ж': 'J',
        'Ч': 'CH', 'Ш': 'SH', 'Щ': 'SH', 'Ъ': '', 'Ь': '', 'Э': 'E', 'Ю': 'YU',
        'Я': 'YA', 'Қ': 'Q', 'Ғ': 'G', 'Ў': 'O', 'Ҳ': 'H', 'X': 'H'
    }
    res = []
    for char in s: res.append(translit_map.get(char, char))
    s = "".join(res)
    s = s.replace('.', '').replace("'", "").replace("`", "").replace('"', "").replace("?", "")
    s = re.sub(r'[^A-Z0-9]', ' ', s)
    words = s.split()
    clean_words = [w for w in words if w not in LEGAL_ENTITIES]
    final_key = " ".join(clean_words).strip()
    return final_key if final_key else " ".join(words)


def clean_number(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ==========================================
# БЕЗОПАСНАЯ РАБОТА СО СЛИТЫМИ ЯЧЕЙКАМИ
# ==========================================
def copy_cell_style_smart(source_cell, target_cell, col_idx):
    if source_cell.has_style:
        try:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            if col_idx <= MAX_TABLE_COL:
                target_cell.border = copy(source_cell.border)
            else:
                from openpyxl.styles import Border
                target_cell.border = Border()
        except Exception:
            pass


def safe_unmerge(ws, coord):
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    for rng in list(ws.merged_cells.ranges):
        if str(rng) == coord:
            ws.merged_cells.remove(rng)
            break
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if (r, c) in ws._cells:
                if isinstance(ws._cells[(r, c)], MergedCell):
                    del ws._cells[(r, c)]


def fix_merged_cells_before_insert(ws, insert_pos):
    broken_merges = []
    merged_ranges = list(ws.merged_cells.ranges)
    for m in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(m.coord)
        if min_row <= insert_pos <= max_row:
            safe_unmerge(ws, m.coord)
            broken_merges.append(m.coord)
    return broken_merges


def shift_merges_down(ws, start_row, shift_amount):
    merges_to_move = []
    for cell_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(cell_range.coord)
        if min_row >= start_row:
            merges_to_move.append((min_col, min_row, max_col, max_row))
            safe_unmerge(ws, cell_range.coord)
    return merges_to_move


def restore_shifted_merges(ws, merges_list, shift_amount):
    for min_col, min_row, max_col, max_row in merges_list:
        new_min_row = min_row + shift_amount
        new_max_row = max_row + shift_amount
        try:
            ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
        except Exception:
            pass


def safe_write(ws, cell_addr, value):
    for rng in list(ws.merged_cells.ranges):
        if cell_addr in rng:
            safe_unmerge(ws, str(rng))
            break
    cell = ws[cell_addr]
    if isinstance(cell, MergedCell) and (cell.row, cell.column) in ws._cells:
        del ws._cells[(cell.row, cell.column)]
    ws[cell_addr] = value


# ==========================================
# ПАРСИНГ И ИНСТЕРТ ДАННЫХ
# ==========================================
def find_our_company_info(ws):
    for r in range(1, 8):
        for c in range(1, 9):
            raw_val = ws.cell(row=r, column=c).value
            if not raw_val: continue
            val = str(raw_val).replace("\xa0", " ").strip()
            match = re.search(r"(\d{20})\s+(.*?)\s+(?:ИНН|inn).*?(\d{9})", val, re.IGNORECASE)
            if match:
                firm_name = match.group(2).strip().strip(".- ")
                firm_inn = match.group(3).strip()
                return firm_name, firm_inn
    return "НАЗВАНИЕ НЕ НАЙДЕНО", ""


def parse_oborotka_row(b_val):
    if not b_val: return None, None
    parts = str(b_val).split('/')
    if len(parts) >= 3:
        inn = parts[1].strip()
        name = "/".join(parts[2:]).strip()
        return inn, name
    return None, None


def check_keywords(text):
    if not text: return False
    t = str(text).lower()
    for kw in KEYWORDS:
        if kw in t: return True
    return False


def collect_data_by_name(wb_ob):
    collected = {}
    our_firm_name, our_firm_inn = "НЕ НАЙДЕНО", ""

    if wb_ob.sheetnames:
        first_ws = wb_ob[wb_ob.sheetnames[0]]
        our_firm_name, our_firm_inn = find_our_company_info(first_ws)

    for ws in wb_ob.worksheets:
        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row or len(row) < 8: continue

            b_val = row[1]
            inn, raw_firm_name = parse_oborotka_row(b_val)
            if not raw_firm_name: continue

            h_val = row[7]
            if not check_keywords(h_val): continue

            debit_val = clean_number(row[5])
            credit_val = clean_number(row[6])

            if debit_val == 0 and credit_val == 0: continue

            unique_key = normalize_firm_name_key(raw_firm_name)
            if not unique_key: continue

            if unique_key not in collected:
                collected[unique_key] = {'display_name': raw_firm_name, 'inn': inn, 'rows': []}

            if len(raw_firm_name) > len(collected[unique_key]['display_name']):
                collected[unique_key]['display_name'] = raw_firm_name

            collected[unique_key]['rows'].append({
                'date': row[0], 'num': row[2], 'desc': h_val,
                'debit': debit_val, 'credit': credit_val
            })

    return collected, our_firm_name, our_firm_inn


def insert_data_into_sheet(ws, start_row, data_rows):
    total_rows = len(data_rows)
    if total_rows == 0: return start_row

    fix_merged_cells_before_insert(ws, start_row)
    saved_merges = shift_merges_down(ws, start_row, total_rows)
    ws.insert_rows(start_row, amount=total_rows)
    restore_shifted_merges(ws, saved_merges, total_rows)

    source_row_idx = start_row + total_rows
    max_col = ws.max_column

    for i in range(total_rows):
        current_r = start_row + i
        for c in range(1, max_col + 1):
            source_cell = ws.cell(row=source_row_idx, column=c)
            target_cell = ws.cell(row=current_r, column=c)
            copy_cell_style_smart(source_cell, target_cell, c)

    current_r = start_row
    for row_data in data_rows:
        ws[f"A{current_r}"] = row_data['num']
        ws[f"B{current_r}"] = row_data['date']
        ws[f"C{current_r}"] = "Финансовая помощь"

        d_val = row_data['debit']
        c_val = row_data['credit']

        if d_val > 0: safe_write(ws, f"D{current_r}", d_val)
        if c_val > 0: safe_write(ws, f"E{current_r}", c_val)

        current_r += 1

    return start_row + total_rows


def apply_formulas(ws, start_row):
    r = start_row
    while r < ws.max_row + 200:
        if not ws[f"A{r}"].value and ws[f"D{r}"].value is None and ws[f"E{r}"].value is None:
            break
        r += 1
    last_data_row = r - 1

    if last_data_row < start_row: return

    for rr in range(start_row, last_data_row + 1):
        safe_write(ws, f"F{rr}", f"=E{rr}")
        safe_write(ws, f"G{rr}", f"=D{rr}")

    sum_row = last_data_row + 2
    safe_write(ws, f"D{sum_row}", f"=SUM(D{start_row}:D{last_data_row})")
    safe_write(ws, f"E{sum_row}", f"=SUM(E{start_row}:E{last_data_row})")
    safe_write(ws, f"F{sum_row}", f"=SUM(F{start_row}:F{last_data_row})")
    safe_write(ws, f"G{sum_row}", f"=SUM(G{start_row}:G{last_data_row})")

    balance_row = sum_row + 1
    safe_write(ws, f"D{balance_row}", f'=IF(D{sum_row}-E{sum_row}>0, D{sum_row}-E{sum_row}, "")')
    safe_write(ws, f"E{balance_row}", f'=IF(E{sum_row}-D{sum_row}>0, E{sum_row}-D{sum_row}, "")')
    safe_write(ws, f"F{balance_row}", f'=IF(F{sum_row}-G{sum_row}>0, F{sum_row}-G{sum_row}, "")')
    safe_write(ws, f"G{balance_row}", f'=IF(G{sum_row}-F{sum_row}>0, G{sum_row}-F{sum_row}, "")')


def copy_sheet_full(wb_target, sheet_name, template_path=TEMPLATE_FILE):
    if not os.path.exists(template_path): return wb_target.create_sheet(title=sheet_name)
    wb_tpl = load_workbook(template_path)
    ws_tpl = wb_tpl.active

    if sheet_name in wb_target.sheetnames:
        ws_new = wb_target[sheet_name]
    else:
        ws_new = wb_target.create_sheet(title=sheet_name)
    ws_new.sheet_view.showGridLines = False

    max_row = ws_tpl.max_row
    max_col = ws_tpl.max_column
    for row in ws_tpl.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell): continue
            new_cell = ws_new.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            copy_cell_style_smart(cell, new_cell, cell.col_idx)

    for merged_range in list(ws_tpl.merged_cells.ranges):
        try:
            ws_new.merge_cells(str(merged_range))
        except Exception:
            pass

    for col_letter, col_dim in ws_tpl.column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width
    for row_idx, row_dim in ws_tpl.row_dimensions.items():
        ws_new.row_dimensions[row_idx].height = row_dim.height

    return ws_new


# ==========================================
# ГЛАВНАЯ ФУНКЦИЯ ДЛЯ БОТА
# ==========================================
def generate_finhelp_acts(oborotka_path):
    """Возвращает (Успех(bool), Сообщение/Путь к файлу)"""
    try:
        if not os.path.exists(TEMPLATE_FILE):
            return False, f"❌ Шаблон {TEMPLATE_FILE} не найден!"

        # ПРОВЕРКА НА 1 ЛИСТ!
        wb_ob = load_workbook(oborotka_path, data_only=True)
        if len(wb_ob.sheetnames) != 1:
            wb_ob.close()
            return False, "❌ Ошибка: В файле Оборотки должен быть строго <b>1 лист</b>!\nПожалуйста, удалите лишние листы и отправьте файл заново."

        collected_data, our_name, our_inn = collect_data_by_name(wb_ob)

        if not collected_data:
            wb_ob.close()
            return False, "⚠️ По ключевым словам ('финанс', 'молияв') ничего не найдено."

        wb_target = Workbook()

        for key, data in collected_data.items():
            partner_name = data['display_name']
            partner_inn = data['inn']
            rows = data['rows']

            clean_name = sanitize_filename_part(key)[:20]
            sheet_title = f"{clean_name}_{partner_inn}"[:31]

            ws_target = copy_sheet_full(wb_target, sheet_title, TEMPLATE_FILE)

            ws_target["B5"] = our_name
            ws_target["C7"] = our_inn

            clean_partner_name = partner_name.replace('""', '"').strip()
            ws_target["E2"] = clean_partner_name
            ws_target["F7"] = partner_inn

            insert_data_into_sheet(ws_target, INSERT_START_ROW, rows)
            apply_formulas(ws_target, INSERT_START_ROW)

        if len(wb_target.sheetnames) > 1 and "Sheet" in wb_target.sheetnames:
            del wb_target["Sheet"]

        now_str = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        output_path = os.path.join(CURRENT_DIR, f"ФинПомощь_{now_str}.xlsx")
        wb_target.save(output_path)
        wb_ob.close()

        return True, output_path

    except Exception as e:
        return False, f"❌ Системная ошибка: {e}"