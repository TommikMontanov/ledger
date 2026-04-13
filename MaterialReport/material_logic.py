import shutil
import pandas as pd
import openpyxl
import os
import re
from copy import copy
from collections import Counter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime

# ================= КОНФИГУРАЦИЯ =================
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(CURRENT_DIR, "Данные.xlsx")

# Константы структуры Excel
START_ROW = 7
HEAD_ROW = 4
PRIHOD_ROW = 3
SUB_HEAD_ROW = 5
BASE_COL = 6  # Столбец F

MONTH_NAMES = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
    7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}


# ================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =================
def get_safe_filename(company_name):
    if not company_name: return "Без названия"
    name = str(company_name).replace('"', '').replace("'", "")
    return re.sub(r'[\\/*?:"<>|]', '_', name).strip()


def break_vertical_merges(ws, start_col, count_cols):
    for i in range(count_cols):
        col_idx = start_col + i
        ranges_to_remove = []
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= col_idx <= rng.max_col:
                if rng.min_row <= 3 and rng.max_row >= 4:
                    ranges_to_remove.append(rng)
        for rng in ranges_to_remove:
            try:
                ws.unmerge_cells(str(rng))
            except:
                if rng in ws.merged_cells: ws.merged_cells.remove(rng)


def reconstruct_headers(ws, start_col_idx, count_pairs):
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=Side('thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_font = Font(name="Calibri", size=8, bold=False)

    for i in range(count_pairs):
        col_q = start_col_idx + (i * 2)
        col_c = col_q + 1

        try:
            ws.unmerge_cells(start_row=HEAD_ROW, start_column=col_q, end_row=HEAD_ROW, end_column=col_c)
        except:
            pass

        ws.merge_cells(start_row=HEAD_ROW, start_column=col_q, end_row=HEAD_ROW, end_column=col_c)

        cell_firm = ws.cell(HEAD_ROW, col_q)
        cell_firm.border = border_all
        cell_firm.alignment = center_align
        ws.cell(HEAD_ROW, col_c).border = border_all

        c_q = ws.cell(SUB_HEAD_ROW, col_q)
        c_q.value = "кол-во"
        c_q.font = sub_font
        c_q.alignment = center_align
        c_q.border = border_all

        c_c = ws.cell(SUB_HEAD_ROW, col_c)
        c_c.value = "сумма"
        c_c.font = sub_font
        c_c.alignment = center_align
        c_c.border = border_all


def expand_prihod_merge(ws, total_col_idx):
    ranges_to_remove = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row == PRIHOD_ROW and rng.max_row == PRIHOD_ROW:
            if rng.max_col >= BASE_COL and rng.min_col <= total_col_idx + 2:
                ranges_to_remove.append(rng)
    for rng in ranges_to_remove:
        try:
            ws.unmerge_cells(str(rng))
        except:
            if rng in ws.merged_cells: ws.merged_cells.remove(rng)

    end_col = total_col_idx - 1
    if end_col > BASE_COL:
        ws.merge_cells(start_row=PRIHOD_ROW, start_column=BASE_COL, end_row=PRIHOD_ROW, end_column=end_col)
        main_cell = ws.cell(PRIHOD_ROW, BASE_COL)
        thin = Side(style='thin')
        main_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        main_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(BASE_COL, end_col + 1):
            ws.cell(PRIHOD_ROW, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fix_total_incoming_header(ws, total_col_idx):
    col_q = total_col_idx
    col_c = total_col_idx + 1
    for r in [3, 4]:
        for c in [col_q, col_c]:
            ranges_to_remove = []
            for rng in ws.merged_cells.ranges:
                if (rng.min_row <= r <= rng.max_row) and (rng.min_col <= c <= rng.max_col):
                    ranges_to_remove.append(rng)
            for rng in ranges_to_remove:
                try:
                    ws.unmerge_cells(str(rng))
                except:
                    if rng in ws.merged_cells: ws.merged_cells.remove(rng)

    ws.merge_cells(start_row=3, start_column=col_q, end_row=4, end_column=col_c)
    main_cell = ws.cell(3, col_q)
    main_cell.value = "Всего приход"
    thin = Side(style='thin')
    border_thick = Border(left=thin, right=thin, top=thin, bottom=thin)
    main_cell.border = border_thick
    main_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    main_cell.font = Font(name="Calibri", size=10, bold=True)
    for r in [3, 4]:
        for c in [col_q, col_c]:
            ws.cell(r, c).border = border_thick

    sub_font = Font(name="Calibri", size=8, bold=False)
    c_q = ws.cell(5, col_q)
    c_q.value = "кол-во"
    c_q.font = sub_font
    c_q.alignment = Alignment(horizontal="center", vertical="center")
    c_q.border = border_thick

    c_c = ws.cell(5, col_c)
    c_c.value = "сумма"
    c_c.font = sub_font
    c_c.alignment = Alignment(horizontal="center", vertical="center")
    c_c.border = border_thick


def apply_final_borders(ws, max_row, max_col):
    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for r in range(START_ROW, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if not cell.border.left.style:
                cell.border = thin_border
    total_row = None
    for r in range(START_ROW, max_row + 2):
        if ws.cell(r, 1).value and "Итого" in str(ws.cell(r, 1).value):
            total_row = r
            break
    if total_row:
        for c in range(1, max_col + 1):
            ws.cell(total_row, c).border = thin_border


def find_total_row_and_col(ws):
    t_row = ws.max_row + 1
    t_col = None
    for r in range(START_ROW, ws.max_row + 100):
        val = ws.cell(r, 1).value
        if val and "Итого" in str(val):
            t_row = r
            break
    for r in [3, 4]:
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val and "Всего приход" in str(val):
                t_col = c
                break
        if t_col: break
    if not t_col: t_col = ws.max_column + 1
    return t_row, t_col


# ================= ПАРСИНГ ДАННЫХ =================
def parse_supplier_data(path):
    if not os.path.exists(path):
        return {}, "", ""
    try:
        df = pd.read_excel(path, header=None)
    except Exception as e:
        return {}, "", ""

    out = {}
    current_supplier = None
    skip_current_block = False
    main_company_name = None
    all_months = []

    COL_ID_DOC = 1
    COL_STATUS = 5
    COL_DATE = 7
    COL_SUPPLIER = 12
    COL_BUYER = 16
    COL_ITEM_NUM = 20
    COL_ITEM_NAME = 21
    COL_UNIT = 23
    COL_QTY = 25
    COL_PRICE = 26
    COL_SUM = 29

    for index, row in df.iterrows():
        r = row.values

        def get_val(idx):
            if idx < len(r) and pd.notna(r[idx]): return str(r[idx]).strip()
            return ""

        val_id = get_val(COL_ID_DOC)

        if main_company_name is None and val_id == '1':
            buyer_val = get_val(COL_BUYER)
            if len(buyer_val) > 1: main_company_name = buyer_val

        if len(r) > COL_DATE and pd.notna(r[COL_DATE]):
            try:
                dt = pd.to_datetime(r[COL_DATE], dayfirst=True, errors='coerce')
                if pd.notna(dt): all_months.append(dt.month)
            except:
                pass

        val_status = get_val(COL_STATUS)
        val_supplier = get_val(COL_SUPPLIER)
        val_u = get_val(COL_ITEM_NUM)

        if val_status or val_u == "Общ.":
            if "Отменен" in val_status:
                skip_current_block = True
            elif val_status:
                skip_current_block = False
            if val_supplier: current_supplier = val_supplier

        if not skip_current_block and val_u.isdigit():
            if not current_supplier: continue
            item_name = get_val(COL_ITEM_NAME)
            unit = get_val(COL_UNIT).lower()

            if "услуг" in get_val(33).lower(): continue

            def clean_num(v):
                if pd.isna(v): return 0.0
                s = str(v).replace(' ', '').replace(',', '.').replace('\xa0', '')
                try:
                    return float(s)
                except:
                    return 0.0

            qty = clean_num(r[COL_QTY]) if len(r) > COL_QTY else 0.0
            price = clean_num(r[COL_PRICE]) if len(r) > COL_PRICE else 0.0
            cost = clean_num(r[COL_SUM]) if len(r) > COL_SUM else 0.0

            if not item_name: item_name = f"Товар стр.{index}"

            items_list = out.setdefault(current_supplier, [])
            items_list.append({"item": item_name, "unit": unit, "quantity": qty, "price": price, "cost": cost})

    m_str = ""
    if all_months:
        common_month = Counter(all_months).most_common(1)[0][0]
        m_str = MONTH_NAMES.get(common_month, "")

    return out, (main_company_name or "Без названия"), m_str


# ================= ГЛАВНАЯ ФУНКЦИЯ ДЛЯ БОТА =================
def generate_material_report(source_file):
    """Принимает путь к загруженному файлу, возвращает (Успех, Путь или Ошибка)"""
    try:
        supp, company_name, report_month = parse_supplier_data(source_file)

        if not supp:
            return False, "❌ Данные не найдены в файле или файл имеет неверный формат."

        if not os.path.exists(TEMPLATE):
            return False, f"❌ Ошибка: Не найден файл шаблона: {TEMPLATE}"

        safe_comp = get_safe_filename(company_name)
        timestamp = datetime.now().strftime("%d%m%H%M")
        output_filename = f"Материальный_отчет_{safe_comp}_{timestamp}.xlsx"
        output_path = os.path.join(CURRENT_DIR, output_filename)

        shutil.copyfile(TEMPLATE, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        if company_name: ws['G1'].value = company_name
        if report_month: ws['J1'].value = report_month

        initial_total_row, total_in_col = find_total_row_and_col(ws)

        current_slots = (total_in_col - BASE_COL) // 2
        needed_slots = len(supp)

        if needed_slots > current_slots:
            diff = needed_slots - current_slots
            cols_to_add = diff * 2
            ws.insert_cols(total_in_col, amount=cols_to_add)
            break_vertical_merges(ws, total_in_col, cols_to_add)
            reconstruct_headers(ws, total_in_col, diff)
            total_in_col += cols_to_add
            expand_prihod_merge(ws, total_in_col)
            fix_total_incoming_header(ws, total_in_col)

        current_row = START_ROW
        while True:
            v1 = ws.cell(current_row, 1).value
            if (not v1) or ("Итого" in str(v1)): break
            current_row += 1

        sorted_suppliers = sorted(supp.keys())

        for i, supplier_name in enumerate(sorted_suppliers):
            items = supp[supplier_name]
            col_idx_q = BASE_COL + (i * 2)
            col_idx_c = col_idx_q + 1

            cell_head = ws.cell(HEAD_ROW, col_idx_q)
            cell_head.value = supplier_name
            cell_head.font = Font(name="Calibri", size=9, bold=True)
            cell_head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for it in items:
                check_total_row, _ = find_total_row_and_col(ws)
                if current_row >= check_total_row:
                    ws.insert_rows(current_row)

                ws.cell(current_row, 1).value = it["item"]
                ws.cell(current_row, 2).value = it["unit"]
                ws.cell(current_row, 2).alignment = Alignment(horizontal="center")
                ws.cell(current_row, 3).value = it["price"]
                ws.cell(current_row, 3).number_format = '#,##0.00'

                c_q = ws.cell(current_row, col_idx_q)
                c_q.value = it["quantity"]
                c_q.number_format = '#,##0.00'
                c_q.alignment = Alignment(horizontal="center")

                c_c = ws.cell(current_row, col_idx_c)
                c_c.value = it["cost"]
                c_c.number_format = '#,##0.00'

                current_row += 1

        final_total_row, _ = find_total_row_and_col(ws)
        apply_final_borders(ws, final_total_row, total_in_col + 5)

        last_data_col_letter = get_column_letter(total_in_col - 1)
        thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        for col in range(BASE_COL, total_in_col):
            col_let = get_column_letter(col)
            cell = ws.cell(final_total_row, col)
            cell.value = f"=SUM({col_let}{START_ROW}:{col_let}{final_total_row - 1})"
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
            cell.border = thin_border

        col_total_qty = total_in_col
        col_total_sum = total_in_col + 1
        let_total_qty = get_column_letter(col_total_qty)
        let_total_sum = get_column_letter(col_total_sum)
        range_header = f"$F$5:${last_data_col_letter}$5"

        for r in range(START_ROW, final_total_row + 1):
            cell_q = ws.cell(r, col_total_qty)
            cell_s = ws.cell(r, col_total_sum)
            cell_q.border = thin_border
            cell_s.border = thin_border
            cell_q.number_format = '#,##0.00'
            cell_s.number_format = '#,##0.00'

            if r == final_total_row:
                cell_q.value = f"=SUM({let_total_qty}{START_ROW}:{let_total_qty}{final_total_row - 1})"
                cell_s.value = f"=SUM({let_total_sum}{START_ROW}:{let_total_sum}{final_total_row - 1})"
                cell_q.font = Font(bold=True)
                cell_s.font = Font(bold=True)
            else:
                range_data = f"$F{r}:${last_data_col_letter}{r}"
                cell_q.value = f'=SUMIF({range_header}, "кол-во", {range_data})'
                cell_s.value = f'=SUMIF({range_header}, "сумма", {range_data})'

        wb.save(output_path)
        return True, output_path

    except Exception as e:
        return False, f"❌ Ошибка обработки материального отчета: {e}"