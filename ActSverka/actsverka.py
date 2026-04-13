import os
import pandas as pd
import csv
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries
from copy import copy

# ---------- Настройки путей ----------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_DB_FILE = os.path.join(CURRENT_DIR, "db_oborotka.xlsx")
TEMPLATE_FILE = os.path.join(CURRENT_DIR, "ActSverkaExample.xlsx")
TEMP_INVOICE_FILE = os.path.join(CURRENT_DIR, "temp_invoices_merged.xlsx")

INSERT_START_ROW = 17
MAX_TABLE_COL = 8


# ==========================================
# 1. ЧИСЛО ПРОПИСЬЮ (СЎМЫ И ТИЙИНЫ)
# ==========================================
def num_to_text_ru(amount):
    """Преобразует число в сумму прописью (сўмы и тийины)"""
    units = (
        ('ноль', 'один', 'два', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь', 'девять'),
        ('десять', 'одиннадцать', 'двенадцать', 'тринадцать', 'четырнадцать', 'пятнадцать', 'шестнадцать', 'семнадцать',
         'восемнадцать', 'девятнадцать'),
        ('ноль', 'десять', 'двадцать', 'тридцать', 'сорок', 'пятьдесят', 'шестьдесят', 'семьдесят', 'восемьдесят',
         'девяносто'),
    )
    hundreds = ('ноль', 'сто', 'двести', 'триста', 'четыреста', 'пятьсот', 'шестьсот', 'семьсот', 'восемьсот',
                'девятьсот')
    thousands = ('тысяча', 'тысячи', 'тысяч')
    millions = ('миллион', 'миллиона', 'миллионов')
    milliards = ('миллиард', 'миллиарда', 'миллиардов')

    def _get_case(n, forms):
        n = abs(int(n)) % 100
        n1 = n % 10
        if 10 < n < 20: return forms[2]
        if 1 < n1 < 5: return forms[1]
        if n1 == 1: return forms[0]
        return forms[2]

    def _three_digits(n, is_female=False):
        res = []
        h, t, u = n // 100, (n % 100) // 10, n % 10
        if h > 0: res.append(hundreds[h])
        if t == 1:
            res.append(units[1][u])
        else:
            if t > 1: res.append(units[2][t])
            if u > 0:
                if is_female:
                    if u == 1:
                        res.append('одна')
                    elif u == 2:
                        res.append('две')
                    else:
                        res.append(units[0][u])
                else:
                    res.append(units[0][u])
        return ' '.join(res)

    amount = clean_number(amount)
    int_part = int(amount)
    frac_part = int(round((amount - int_part) * 100))

    parts = []
    mrd = (int_part // 1000000000) % 1000
    if mrd > 0:
        parts.append(_three_digits(mrd))
        parts.append(_get_case(mrd, milliards))
    mln = (int_part // 1000000) % 1000
    if mln > 0:
        parts.append(_three_digits(mln))
        parts.append(_get_case(mln, millions))
    th = (int_part // 1000) % 1000
    if th > 0:
        parts.append(_three_digits(th, is_female=True))
        parts.append(_get_case(th, thousands))
    sm = int_part % 1000
    if sm > 0 or not parts:
        parts.append(_three_digits(sm))

    som_forms = ('сўм', 'сўма', 'сўмов')
    parts.append(_get_case(int_part, som_forms))

    tiyin_forms = ('тийин', 'тийина', 'тийин')
    res_str = ' '.join(parts).capitalize()
    return f"{res_str} {frac_part:02d} {tiyin_forms[2]}"


# ==========================================
# 2. УТИЛИТЫ И ОБРАБОТКА ЧИСЕЛ
# ==========================================
def sanitize_filename_part(s: str) -> str:
    if s is None: return "UNKNOWN"
    t = str(s).strip()
    t = re.sub(r'[\\/*?:"<>|\n\r\t]', "", t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t or "UNKNOWN"


def clean_number(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def apply_currency_format(cell):
    """Применяет бухгалтерский формат с пробелами и запятой"""
    cell.number_format = '#,##0.00'


# ==========================================
# 3. ФОРМУЛЫ И ИТОГИ (ВКЛЮЧАЯ E223)
# ==========================================
def apply_formulas(ws, start_row):
    r = start_row
    while r < ws.max_row + 200:
        val = ws[f"A{r}"].value
        if not val and ws[f"D{r}"].value is None and ws[f"E{r}"].value is None:
            break
        r += 1
    last_data_row = r - 1

    if last_data_row < start_row: return 0.0

    sum_d, sum_e = 0.0, 0.0
    for rr in range(start_row, last_data_row + 1):
        ws[f"F{rr}"] = f"=E{rr}"
        ws[f"G{rr}"] = f"=D{rr}"

        for col in "DEFG":
            apply_currency_format(ws[f"{col}{rr}"])

        sum_d += clean_number(ws[f"D{rr}"].value)
        sum_e += clean_number(ws[f"E{rr}"].value)

    sum_f = sum_e
    sum_g = sum_d

    # СТРОКА: Всего Обороты
    sum_row = last_data_row + 2
    for col in "DEFG":
        ws[f"{col}{sum_row}"] = f"=SUM({col}{start_row}:{col}{last_data_row})"
        apply_currency_format(ws[f"{col}{sum_row}"])

    # СТРОКА: Сальдо (Баланс)
    balance_row = sum_row + 1
    diff_1 = sum_d - sum_e

    ws[f"D{balance_row}"] = None
    ws[f"E{balance_row}"] = None
    ws[f"F{balance_row}"] = None
    ws[f"G{balance_row}"] = None

    if diff_1 < 0:
        ws[f"D{balance_row}"] = f"=E{sum_row}-D{sum_row}"
        apply_currency_format(ws[f"D{balance_row}"])
    else:
        ws[f"E{balance_row}"] = f"=D{sum_row}-E{sum_row}"
        apply_currency_format(ws[f"E{balance_row}"])

    diff_2 = sum_f - sum_g
    if diff_2 > 0:
        ws[f"G{balance_row}"] = f"=F{sum_row}-G{sum_row}"
        apply_currency_format(ws[f"G{balance_row}"])
    else:
        ws[f"F{balance_row}"] = f"=G{sum_row}-F{sum_row}"
        apply_currency_format(ws[f"F{balance_row}"])

    # --- ЗАМЕНА СТАРЫХ ДАННЫХ В ПОДВАЛЕ (СУММА ПРОПИСЬЮ И E223) ---
    final_balance_abs = abs(diff_1)
    text_sum = num_to_text_ru(final_balance_abs)

    for r_search in range(balance_row, ws.max_row + 15):
        for c_search in range(1, 12):
            cell = ws.cell(row=r_search, column=c_search)
            val_str = str(cell.value or "").lower()

            # 1. Замена текста (Сўмы прописью)
            if "сўм" in val_str or "тийин" in val_str or "сумов" in val_str:
                cell.value = text_sum

            # 2. Замена нуля на финальную сумму (тот самый старый 0 в колонке E или соседних)
            if isinstance(cell.value, (int, float)) and cell.value == 0:
                if c_search in [4, 5, 6, 7]:  # Колонки D, E, F, G
                    cell.value = final_balance_abs
                    apply_currency_format(cell)

    return final_balance_abs


# ==========================================
# 4. РАБОТА С ШАБЛОНАМИ И СТИЛЯМИ
# ==========================================
def copy_cell_style_smart(source_cell, target_cell, col_idx):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        if col_idx <= MAX_TABLE_COL:
            target_cell.border = copy(source_cell.border)


def fix_merged_cells_before_insert(ws, insert_pos):
    broken_merges = []
    merged_ranges = list(ws.merged_cells.ranges)
    for m in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(m.coord)
        if min_row <= insert_pos <= max_row:
            ws.unmerge_cells(m.coord)
            broken_merges.append(m.coord)
    return broken_merges


def shift_merges_down(ws, start_row, shift_amount):
    merges_to_move = []
    for cell_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(cell_range.coord)
        if min_row >= start_row:
            merges_to_move.append((min_col, min_row, max_col, max_row))
            ws.unmerge_cells(cell_range.coord)
    return merges_to_move


def restore_shifted_merges(ws, merges_list, shift_amount):
    for min_col, min_row, max_col, max_row in merges_list:
        new_min_row = min_row + shift_amount
        new_max_row = max_row + shift_amount
        try:
            ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
        except:
            pass


def insert_all_data_bulk(ws, start_row, payments, invoices, code):
    total_rows = len(payments) + len(invoices)
    if total_rows == 0: return start_row

    fix_merged_cells_before_insert(ws, start_row)
    saved_merges = shift_merges_down(ws, start_row, total_rows)
    ws.insert_rows(start_row, amount=total_rows)
    restore_shifted_merges(ws, saved_merges, total_rows)

    source_row_idx = start_row + total_rows
    max_col = ws.max_column

    for i in range(total_rows):
        current_r = start_row + i
        for c in range(1, max_col + 1):
            source_cell = ws.cell(row=source_row_idx, column=c)
            target_cell = ws.cell(row=current_r, column=c)
            copy_cell_style_smart(source_cell, target_cell, c)

    current_r = start_row

    for doc_num, date_val, amount in payments:
        ws[f"A{current_r}"] = doc_num
        ws[f"B{current_r}"] = date_val
        ws[f"C{current_r}"] = "Платежное поручение"
        val_num = clean_number(amount)
        col = "D" if code == "6010" else "E"
        ws[f"{col}{current_r}"] = val_num
        apply_currency_format(ws[f"{col}{current_r}"])
        current_r += 1

    for nomer, date, summa, status in invoices:
        ws[f"A{current_r}"] = nomer
        ws[f"B{current_r}"] = date
        ws[f"C{current_r}"] = "Счет Фактура"
        val_num = clean_number(summa)
        col = "E" if code == "6010" else "D"
        ws[f"{col}{current_r}"] = val_num
        ws[f"H{current_r}"] = status
        apply_currency_format(ws[f"{col}{current_r}"])
        current_r += 1

    return start_row + total_rows


def copy_sheet_full(wb_target, sheet_name, template_path):
    if not os.path.exists(template_path): return wb_target.create_sheet(title=sheet_name)
    wb_tpl = load_workbook(template_path)
    ws_tpl = wb_tpl.active

    if sheet_name in wb_target.sheetnames:
        ws_new = wb_target[sheet_name]
    else:
        ws_new = wb_target.create_sheet(title=sheet_name)
    ws_new.sheet_view.showGridLines = False

    max_row = ws_tpl.max_row
    max_col = ws_tpl.max_column
    for row in ws_tpl.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell): continue
            new_cell = ws_new.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            copy_cell_style_smart(cell, new_cell, cell.col_idx)

    for merged_range in list(ws_tpl.merged_cells.ranges):
        try:
            ws_new.merge_cells(str(merged_range))
        except:
            pass

    for col_letter, col_dim in ws_tpl.column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width
    for row_idx, row_dim in ws_tpl.row_dimensions.items():
        ws_new.row_dimensions[row_idx].height = row_dim.height

    try:
        if ws_tpl.freeze_panes: ws_new.freeze_panes = ws_tpl.freeze_panes
    except:
        pass
    return ws_new


# ==========================================
# 5. СБОР ДАННЫХ ИЗ ФАЙЛОВ
# ==========================================
def collect_oborotka_rows(ws_ob, inn_input, amount_index):
    rows = []
    inn_input = str(inn_input).strip() if inn_input is not None else ""
    if not inn_input: return rows
    for r in ws_ob.iter_rows(min_row=8, values_only=True):
        if not any(r): continue
        inn_b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        inn_h = str(r[7]).strip() if len(r) > 7 and r[7] is not None else ""
        if inn_input and (inn_input in inn_b or inn_input in inn_h):
            raw_amount = r[amount_index] if len(r) > amount_index else None
            amount = clean_number(raw_amount)
            if amount == 0: continue
            date_val = r[0] if len(r) > 0 else None
            doc_num = r[2] if len(r) > 2 else None
            rows.append((doc_num, date_val, amount))
    return rows


def collect_schet_rows_from_sheet(ws_schet):
    rows = []
    for r in ws_schet.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        d_col = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        raw_summa = r[14] if len(r) > 14 else None
        summa = clean_number(raw_summa)
        status = r[2] if len(r) > 2 else None
        parts = re.split(r'\s+от\s+', d_col, flags=re.IGNORECASE)
        if len(parts) >= 2:
            nomer, date = parts[0].strip(), parts[1].strip()
        else:
            nomer, date = d_col, ""
        rows.append((nomer, date, summa, status))
    return rows


def detect_encoding(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            f.read()
        return "utf-8"
    except:
        return "cp1251"


def get_sheet_name_from_csv(file_path, encoding):
    with open(file_path, "r", encoding=encoding, newline="") as f:
        reader = list(csv.reader(f, delimiter=";"))
        if len(reader) < 2: return os.path.basename(file_path)[:31]
        for col_idx in [7, 11]:
            if len(reader[0]) > col_idx and 'ПРОДАВЕЦ' in str(reader[0][col_idx]).upper():
                val = str(reader[1][col_idx]).strip()
                return re.sub(r'[\\/*?:"<>|]', "", val)[:31]
        return os.path.basename(file_path)[:31]


def merge_csv_to_excel(csv_files):
    if os.path.exists(TEMP_INVOICE_FILE): os.remove(TEMP_INVOICE_FILE)
    for file_path in csv_files:
        enc = detect_encoding(file_path)
        sh_name = get_sheet_name_from_csv(file_path, enc)
        df = pd.read_csv(file_path, encoding=enc, sep=";")
        if not os.path.exists(TEMP_INVOICE_FILE):
            df.to_excel(TEMP_INVOICE_FILE, sheet_name=sh_name, index=False)
        else:
            with pd.ExcelWriter(TEMP_INVOICE_FILE, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
                df.to_excel(writer, sheet_name=sh_name, index=False)


# ==========================================
# 6. ГЛАВНЫЕ ФУНКЦИИ ДЛЯ БОТА
# ==========================================
def update_master_oborotka(new_file_path):
    """ПОЛНАЯ ЗАМЕНА БАЗЫ ДАННЫХ: удаляем старую, ставим новую"""
    try:
        if os.path.exists(MASTER_DB_FILE):
            os.remove(MASTER_DB_FILE)  # Удаляем старую базу без следа

        shutil.move(new_file_path, MASTER_DB_FILE)  # Перемещаем новую на её место

        return True, "✅ Старая база удалена, новая база успешно загружена!"
    except Exception as e:
        return False, f"❌ Ошибка при замене БД: {e}"


def process_reconciliation_acts(csv_file_paths, code):
    merge_csv_to_excel(csv_file_paths)
    if not os.path.exists(MASTER_DB_FILE) or not os.path.exists(TEMPLATE_FILE): return []

    wb_ob = load_workbook(MASTER_DB_FILE, data_only=True)
    wb_schet_all = load_workbook(TEMP_INVOICE_FILE, data_only=True)
    wb_target = Workbook()

    amount_idx = 5 if code == "6010" else 6
    generated_files = []

    for sch_name in wb_schet_all.sheetnames:
        ws_schet = wb_schet_all[sch_name]

        inn_for_ob_sheet = ws_schet["G2"].value if code == "4010" else ws_schet["K2"].value
        inn_for_payments = ws_schet["K2"].value if code == "4010" else ws_schet["G2"].value

        if not inn_for_ob_sheet or str(inn_for_ob_sheet).strip() == "": continue

        inn_ob_search = str(inn_for_ob_sheet).strip()
        inn_payment_search = str(inn_for_payments).strip() if inn_for_payments else ""

        found_ws_ob = None
        for ws in wb_ob.worksheets:
            try:
                a5 = str(ws["A5"].value or "")
            except:
                a5 = ""
            if inn_ob_search in a5:
                found_ws_ob = ws
                break

        if not found_ws_ob: continue

        if code == "6010":
            inn_cell_val = ws_schet["G2"].value
            firm_cell_val = ws_schet["H2"].value
        else:
            inn_cell_val = ws_schet["K2"].value
            firm_cell_val = ws_schet["L2"].value

        firm_name = str(firm_cell_val).strip() if firm_cell_val is not None else f"Фирма_{sch_name}"
        inn_value = str(inn_cell_val).strip() if inn_cell_val is not None else ""
        sheet_name = sanitize_filename_part(firm_name)[:31]

        if sheet_name in wb_target.sheetnames:
            ws_target = wb_target[sheet_name]
        else:
            ws_target = copy_sheet_full(wb_target, sheet_name, TEMPLATE_FILE)

        firm_clean_e2 = firm_name.replace('""', '"').replace('«', '').replace('»', '').replace('"', '').strip()
        formatted_firm = f'{inn_value} "{firm_clean_e2}"' if firm_clean_e2 else inn_value
        ws_target["E2"] = formatted_firm
        ws_target["F7"] = inn_value

        if code == "4010":
            inn_b5 = str(ws_schet["G2"].value or "").strip()
            firm_b5 = str(ws_schet["H2"].value or "").strip()
        else:
            inn_b5 = str(ws_schet["K2"].value or "").strip()
            firm_b5 = str(ws_schet["L2"].value or "").strip()

        firm_b5 = firm_b5.replace('"', '').replace('«', '').replace('»', '').strip()
        ws_target["C7"] = inn_b5
        ws_target["B5"] = f'{inn_b5} ООО «{firm_b5}»' if inn_b5 else f'ООО «{firm_b5}»'

        if code == "4010":
            firm_for_b10 = str(ws_schet["H2"].value or "").replace('"', '').replace('«', '').replace('»', '').strip()
        else:
            firm_for_b10 = str(ws_schet["L2"].value or "").replace('"', '').replace('«', '').replace('»', '').strip()

        ws_target["B10"] = f'Мы, нижеподписавшиеся, ООО «{firm_for_b10}», в лице Директора ......., с одной стороны.'

        payments = collect_oborotka_rows(found_ws_ob, inn_payment_search, amount_idx)
        invoices = collect_schet_rows_from_sheet(ws_schet)

        insert_all_data_bulk(ws_target, INSERT_START_ROW, payments, invoices, code)
        apply_formulas(ws_target, INSERT_START_ROW)

    if "Sheet" in wb_target.sheetnames: del wb_target["Sheet"]

    if wb_target.sheetnames:
        now_str = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        output_path = os.path.join(CURRENT_DIR, f"{code}-Акт_Сверка_{now_str}.xlsx")
        wb_target.save(output_path)
        generated_files.append(output_path)

    wb_schet_all.close()
    wb_ob.close()
    if os.path.exists(TEMP_INVOICE_FILE): os.remove(TEMP_INVOICE_FILE)

    return generated_files