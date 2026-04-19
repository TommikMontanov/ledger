import re
import os
import copy
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Константы
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(CURRENT_DIR, "example_4010.xlsx")

months_order = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']


# --- БЛОК 1: АВТОМАТИЧЕСКАЯ КЛАССИФИКАЦИЯ ФАЙЛОВ (FSM) ---

def classify_uploaded_files(file_paths):
    """
    Интеллектуальное распределение файлов по ролям.
    Использует поиск по маскам в названиях листов и контенте.
    """
    classified = {'summary': None, 'registry': None, 'source': None}

    # Ключевые слова для поиска
    MARKERS = {
        'registry': ["реализ", "товар", "реестр", "9030"],
        'source': ["дебет", "кредит", "счет", "4010"],
        'summary': ["период", "сальдо", "обороты", "осв"]
    }

    for path in file_paths:
        if not path or not os.path.exists(path): continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)

            # Сначала проверяем названия листов (самый быстрый способ)
            sheet_names_joined = " ".join(wb.sheetnames).lower()

            # 1. Приоритет №1: Реестр (ищем корни слов)
            if any(m in sheet_names_joined for m in MARKERS['registry']):
                classified['registry'] = path
                wb.close()
                continue

            # 2. Если не реестр, лезем внутрь активного листа
            ws = wb.active
            header_content = ""
            for r in range(1, 12):  # Читаем чуть глубже, до 12 строки
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if val: header_content += str(val).lower() + " "

            # 3. Приоритет №2: Исходник (Сводка по счетам)
            # Если видим Дебет + Кредит + 4010
            if "дебет" in header_content and "кредит" in header_content:
                classified['source'] = path

            # 4. Приоритет №3: Общая Сводка (ОСВ)
            # Если есть даты (ДД.ММ.ГГГГ)
            elif re.search(r'\d{2}\.\d{2}\.\d{4}', header_content):
                classified['summary'] = path

            wb.close()
        except Exception as e:
            print(f"Ошибка анализа {path}: {e}")

    return classified


# --- БЛОК 2: ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ОБРАБОТКИ ---

def get_ident_key(text):
    if not text: return None
    s = str(text).strip()
    if '/' in s:
        parts = s.split('/')
        candidate = parts[0].strip()
        if re.fullmatch(r'\d+', candidate): return candidate
    matches = re.findall(r'\b\d{9,14}\b', s)
    if matches: return matches[-1]
    clean_s = re.sub(r'\s+', ' ', s).strip().upper()
    return clean_s if clean_s else None


def normalize_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try:
        return float(str(val).replace(',', '.').replace(' ', '').replace('\xa0', ''))
    except:
        return 0.0


def find_header_in_column(ws, col_letter, search_text):
    for r in range(1, 100):
        val = ws[f"{col_letter}{r}"].value
        if val and search_text.lower() in str(val).lower(): return r
    return 0


def get_data_start_row(ws):
    r_name = find_header_in_column(ws, "A", "Наименование")
    r_deb = find_header_in_column(ws, "C", "Дебет")
    r_cred = find_header_in_column(ws, "D", "Кредит")
    bottom_header_row = max(r_name, r_deb, r_cred)
    return bottom_header_row + 1 if bottom_header_row != 0 else 5


def find_itogo_row(ws):
    for r in range(1, ws.max_row + 1):
        val = ws[f"A{r}"].value
        if val and isinstance(val, str) and "ИТОГО" in val.upper(): return r
    return ws.max_row + 1


def copy_row_style_and_formulas(ws, source_row_idx, target_row_idx):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=source_row_idx, column=col)
        dst_cell = ws.cell(row=target_row_idx, column=col)
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.alignment = copy.copy(src_cell.alignment)
        val = src_cell.value
        if isinstance(val, str) and val.startswith('='):
            try:
                dst_cell.value = Translator(val, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
            except:
                dst_cell.value = None
        else:
            dst_cell.value = None


def find_or_create_row_for_inn(ws, ident_key, start_row_data):
    itogo_row = find_itogo_row(ws)
    for r in range(start_row_data, itogo_row):
        a_val, b_val = ws[f"A{r}"].value, ws[f"B{r}"].value
        if get_ident_key(a_val) == ident_key or get_ident_key(b_val) == ident_key: return r

    # Если не нашли, вставляем новую строку перед ИТОГО
    ws.insert_rows(itogo_row)
    if itogo_row > start_row_data:
        copy_row_style_and_formulas(ws, itogo_row - 1, itogo_row)
    return itogo_row


def update_itogo_formulas(ws):
    itogo_row = find_itogo_row(ws)
    if itogo_row > ws.max_row: return
    pattern = r'((?:\$?[A-Za-z]+)(?:\$?\d+)):(\$?[A-Za-z]+)(\$?)(\d+)'
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=itogo_row, column=col)
        val = cell.value
        if val and isinstance(val, str) and val.startswith('='):
            new_val = re.sub(pattern, lambda m: f"{m.group(1)}:{m.group(2)}{m.group(3)}{itogo_row - 1}", val)
            if new_val != val: cell.value = new_val


# --- БЛОК 3: ОСНОВНАЯ ГЕНЕРАЦИЯ ---

def generate_svodka_4010(file_paths_list, start_month_idx, is_saldo_transferred):
    try:
        # ШАГ 1: Распознаем файлы через обновленный FSM
        files_dict = classify_uploaded_files(file_paths_list)

        f_src = files_dict['source']
        f_sum = files_dict['summary']
        f_reg = files_dict['registry']

        # Если реестр все еще не найден, попробуем взять методом исключения
        # (если 2 других файла опознаны, то третий — точно реестр)
        if not f_reg:
            all_files = set(file_paths_list)
            identified = {f_src, f_sum}
            remaining = list(all_files - identified)
            if len(remaining) == 1:
                f_reg = remaining[0]

        if not all([f_src, f_sum, f_reg]):
            return False, (f"❌ Ошибка: Не все файлы распознаны.\n"
                           f"• Исходник: {'✅' if f_src else '❌'}\n"
                           f"• Сводка: {'✅' if f_sum else '❌'}\n"
                           f"• Реестр: {'✅' if f_reg else '❌'}\n\n"
                           f"Проверь, что в Реестре есть слова 'реализация' или 'товар'.")

        # ШАГ 2: Работа с Реестром (9030)
        wb_reg = load_workbook(f_reg, data_only=True)
        # Ищем нужный лист не по точному имени, а по вхождению слова
        ws_reg = None
        for s_name in wb_reg.sheetnames:
            if any(k in s_name.lower() for k in ["реализ", "товар", "9030"]):
                ws_reg = wb_reg[s_name]
                break

        if not ws_reg: ws_reg = wb_reg.active  # Fallback на первый лист

        # 3. Читаем Сводку (4010)
        wb_summary = load_workbook(f_sum, data_only=True)
        company_name = wb_summary[wb_summary.sheetnames[0]]["A2"].value

        # Определяем год и доступные месяцы
        year = str(datetime.now().year)
        available_months = {start_month_idx}

        # Анализируем даты в реестре для определения месяцев
        for r in range(1, ws_reg.max_row + 1):
            val = ws_reg[f"E{r}"].value
            if isinstance(val, datetime):
                year = str(val.year)
                available_months.add(val.month - 1)
            elif val and re.search(r'\d{2}\.\d{2}\.(\d{4})', str(val)):
                m = re.search(r'\d{2}\.(\d{2})\.(\d{4})', str(val))
                available_months.add(int(m.group(1)) - 1)
                year = m.group(2)

        months_to_process = [months_order[i] for i in sorted(list(available_months)) if i >= start_month_idx]

        # 4. Читаем Исходник (Сальдо)
        wb_src_file = load_workbook(f_src, data_only=True)
        ws_src = wb_src_file.active
        start_row_src = get_data_start_row(ws_src)

        # Определяем колонки сальдо в исходнике
        col_c_idx, col_d_idx = 3, 4  # По умолчанию C, D
        if not is_saldo_transferred:
            for c in range(15, 22):  # Ищем в районе O-U
                h = str(ws_src.cell(row=1, column=c).value or "").upper() + str(
                    ws_src.cell(row=2, column=c).value or "").upper()
                if "ДЕБЕТ" in h: col_c_idx = c
                if "КРЕДИТ" in h: col_d_idx = c

        base_data, seen_inns = [], set()
        for r in range(start_row_src, find_itogo_row(ws_src)):
            val_a, val_b = ws_src[f"A{r}"].value, ws_src[f"B{r}"].value
            if not val_a or "ИТОГО" in str(val_a).upper(): continue
            key = get_ident_key(val_a) or get_ident_key(val_b)
            base_data.append({
                'key': key, 'name_a': val_a, 'name_b': val_b,
                's_c': normalize_float(ws_src.cell(row=r, column=col_c_idx).value),
                's_d': normalize_float(ws_src.cell(row=r, column=col_d_idx).value)
            })
            if key: seen_inns.add(key)
        wb_src_file.close()

        # 5. Сбор данных по месяцам
        summary_map = defaultdict(lambda: defaultdict(list))
        registry_map = defaultdict(lambda: defaultdict(list))
        names_lookup = {}

        # Из Сводки (4010)
        for m_name in months_to_process:
            # Ищем нужный лист в Сводке
            ws_sum_m = next((wb_summary[s] for s in wb_summary.sheetnames if m_name.lower() in s.lower()), None)
            if ws_sum_m:
                for r in range(4, ws_sum_m.max_row + 1):
                    a, b = ws_sum_m[f"A{r}"].value, ws_sum_m[f"B{r}"].value
                    if not a or "ИТОГО" in str(a).upper(): continue
                    key = get_ident_key(a) or get_ident_key(b)
                    val = normalize_float(ws_sum_m[f"E{r}"].value)  # Предполагаем 4010 в E
                    if key:
                        summary_map[m_name][key].append(val)
                        names_lookup[key] = (str(a), str(b or ""))

        # Из Реестра (9030)
        for r in range(1, ws_reg.max_row + 1):
            dt = ws_reg[f"E{r}"].value
            m_name = None
            if isinstance(dt, datetime):
                m_name = months_order[dt.month - 1]
            elif isinstance(dt, str) and re.search(r'\.\d{2}\.', dt):
                m_idx = int(dt.split('.')[1]) - 1
                m_name = months_order[m_idx]

            if m_name in months_to_process:
                firm, inn, val = ws_reg[f"B{r}"].value, ws_reg[f"C{r}"].value, ws_reg[f"F{r}"].value
                key = get_ident_key(firm) or get_ident_key(inn)
                if key:
                    registry_map[m_name][key].append(normalize_float(val))
                    if key not in names_lookup: names_lookup[key] = (str(firm), str(inn))

        # 6. Формирование финального файла
        wb_out = load_workbook(TEMPLATE_FILE)
        tmpl = wb_out.active
        start_row_tmpl = get_data_start_row(tmpl)

        row_mapping = defaultdict(dict)  # Для хранения номеров строк по листам
        current_entities = list(base_data)

        for m_idx, m_name in enumerate(months_to_process):
            ws = wb_out.copy_worksheet(tmpl)
            ws.title = f"{m_name} {year}"
            ws["J1"], ws["M1"], ws["D1"] = m_name, year, company_name

            # Добавляем новые фирмы, появившиеся в этом месяце
            month_keys = set(summary_map[m_name].keys()) | set(registry_map[m_name].keys())
            for mk in month_keys:
                if mk not in seen_inns:
                    seen_inns.add(mk)
                    n_a, n_b = names_lookup.get(mk, (f"Новая фирма {mk}", ""))
                    current_entities.append({'key': mk, 'name_a': n_a, 'name_b': n_b, 's_c': 0, 's_d': 0})

            # Заполняем строки
            for ent in current_entities:
                target_r = find_or_create_row_for_inn(ws, ent['key'], start_row_tmpl)
                row_mapping[m_name][ent['key']] = target_r

                ws[f"A{target_r}"] = ent['name_a']
                ws[f"B{target_r}"] = ent['name_b']
                ws[f"A{target_r}"].alignment = Alignment(wrap_text=True)

                # Сальдо
                if m_idx == 0:
                    ws[f"C{target_r}"], ws[f"D{target_r}"] = ent['s_c'], ent['s_d']
                else:
                    prev_m = months_to_process[m_idx - 1]
                    prev_r = row_mapping[prev_m].get(ent['key'])
                    if prev_r:
                        ws[f"C{target_r}"] = f"='{prev_m} {year}'!O{prev_r}"
                        ws[f"D{target_r}"] = f"='{prev_m} {year}'!P{prev_r}"

                # Обороты (формулы сумм)
                if ent['key'] in summary_map[m_name]:
                    ws[f"F{target_r}"] = "=" + "+".join(map(str, summary_map[m_name][ent['key']]))
                if ent['key'] in registry_map[m_name]:
                    ws[f"J{target_r}"] = "=" + "+".join(map(str, registry_map[m_name][ent['key']]))

            update_itogo_formulas(ws)

        wb_out.remove(tmpl)
        safe_name = re.sub(r'[\\/*?:"<>|]', "", str(company_name or "Result"))
        out_path = os.path.join(CURRENT_DIR, f"Сводка_4010_{safe_name}.xlsx")
        wb_out.save(out_path)

        # Закрытие всех книг
        wb_summary.close();
        wb_reg.close();
        wb_out.close()
        return True, out_path

    except Exception as e:
        import traceback
        return False, f"❌ Ошибка: {str(e)}\n{traceback.format_exc()}"