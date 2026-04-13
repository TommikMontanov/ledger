import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime


MONTH_NAMES = {
    '01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
    '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
    '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'
}

def to_float_safe(value):
    try:
        if pd.isna(value): return 0.0
        if isinstance(value, (int, float)): return float(value)
        value = str(value).replace(' ', '').replace(',', '.')
        if not value: return 0.0
        return float(value)
    except:
        return 0.0

# ==========================================
# 1. Чтение данных с конкретного листа
# ==========================================
def read_turnover_sheet(file_path, sheet_name, target_month):
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        account_map = {
            "20208000105160429001": "FIST STEP TO FUTURE MCHJ",
            "20208000005461593001": 'ООО "PRO LED TIME"',
            "20208000900342043001": 'XK "MEGA PRINT PLUS"',
            "20208000205542112001": "NUR FAYZ REKLAMA MCHJ",
            "20208000605331536001": "MCHJ SHOXSULTON GOLD KOMPLEKS",
            "20208000400899910001": "ООО NURAFSHON CITY REKLAMA",
            "20208000600926006001": "МЧЖ VENTILYATSIYA BUILDING",
            "16401000000418642001": "МЧЖ MUZAFFAR FARZONA",
            "20208000400418642001": "МЧЖ MUZAFFAR FARZONA"
        }
        SPECIAL_FORMAT_ACCOUNTS = ["20208000600926006001"]

        company_name = "Не указано"
        period = "Не указан"
        is_target_month = False

        for i in range(min(50, len(df_raw))):
            row_str = ' '.join(df_raw.iloc[i].dropna().astype(str)).lower()
            if period == "Не указан":
                if "справка о работе" in row_str or "сведения о работе" in row_str:
                    match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})\s*(?:по|-)\s*(\d{2})\.(\d{2})\.(\d{4})', row_str)
                    if match:
                        start_d, start_m, start_y, end_d, end_m, end_y = match.groups()
                        period = f"{start_d}.{start_m}.{start_y} - {end_d}.{end_m}.{end_y}"
                        if start_m == target_month:
                            is_target_month = True
                        break

        if not is_target_month: return None, None, None, None

        valid_header_idx = None
        start_balance = 0.0
        temp_balance = 0.0
        last_seen_header_idx = 0
        found_account_number = None

        for i in range(len(df_raw)):
            row_vals = df_raw.iloc[i].dropna().astype(str)
            row_str = ' '.join(row_vals).lower().replace('c', 'с')

            if 'дата' in row_str and ('счет' in row_str or 'инн' in row_str or 'наименование' in row_str):
                last_seen_header_idx = i

            if "статок на начало" in row_str:
                after_keyword = row_str.split("статок на начало")[1]
                match = re.search(r'(\d[\d\s]*(?:[.,]\d+)?)', after_keyword)
                if match:
                    num_str = match.group(1).replace(' ', '').replace(',', '.')
                    try: temp_balance = float(num_str)
                    except ValueError: pass

            if 'итоговый оборот' in row_str or 'итого' in row_str:
                has_sums = False
                for val in row_vals:
                    val_clean = val.replace(' ', '').replace(',', '.')
                    try:
                        if float(val_clean) > 0:
                            has_sums = True
                            break
                    except ValueError: pass

                if has_sums:
                    start_balance = temp_balance
                    valid_header_idx = last_seen_header_idx
                    for j in range(max(0, valid_header_idx - 10), valid_header_idx):
                        j_str = ' '.join(df_raw.iloc[j].dropna().astype(str))
                        for acc, firm in account_map.items():
                            if acc in j_str:
                                company_name = firm
                                found_account_number = acc
                                break
                        if found_account_number: break
                    break

        if valid_header_idx is None: return None, None, None, None

        if company_name == "Не указано":
            for j in range(max(0, valid_header_idx - 10), valid_header_idx):
                j_str = ' '.join(df_raw.iloc[j].dropna().astype(str))
                if 'ООО' in j_str:
                    match = re.search(r'ООО\s*"[^"]+"', j_str)
                    if match:
                        company_name = match.group(0)
                        break

        if found_account_number in SPECIAL_FORMAT_ACCOUNTS:
            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=valid_header_idx + 1, header=None, usecols=[0, 1, 2, 3, 6, 7, 8])
            df.columns = ['Дата', 'Счет_raw', 'Название_ИНН_raw', '№ док', 'Оборот Дебет', 'Оборот Кредит', 'Назначение платежа']
            cutoff_index = None
            for i in range(len(df)):
                val_a = str(df.iloc[i, 0]).strip().lower()
                if 'итоговый оборот' in val_a or 'итого' in val_a:
                    cutoff_index = i
                    break
            if cutoff_index is not None: df = df.iloc[:cutoff_index]
            df = df[df['Дата'].notna()]

            def parse_details(acc_val, name_inn_val):
                acc_clean = str(acc_val).strip()
                raw_str = str(name_inn_val).strip()
                inn_match = re.search(r'(\d{9,})$', raw_str)
                if not inn_match: inn_match = re.search(r'(\d+)$', raw_str)
                if inn_match:
                    inn = inn_match.group(1)
                    name = raw_str[:inn_match.start()].strip()
                else:
                    inn = "000000000"
                    name = raw_str
                return f"{acc_clean}/{inn}/{name}"

            df['Cчет/ИНН'] = df.apply(lambda x: parse_details(x['Счет_raw'], x['Название_ИНН_raw']), axis=1)
            df['Дата'] = pd.to_datetime(df['Дата'], dayfirst=True, errors='coerce')
            return df.dropna(subset=['Дата']), company_name, period, start_balance

        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=valid_header_idx)
            df.columns = [str(col).strip() for col in df.columns]
            df.columns = df.columns.str.replace('С', 'C', regex=False)
            if 'Дата' not in df.columns:
                possible_date_cols = [c for c in df.columns if 'дата' in str(c).lower()]
                if possible_date_cols: df.rename(columns={possible_date_cols[0]: 'Дата'}, inplace=True)
                else: raise KeyError("Столбец 'Дата' не найден.")

            cutoff_index = None
            for i in range(len(df)):
                val_a = str(df['Дата'].iloc[i]).strip().lower()
                if 'итоговый оборот' in val_a or 'итого' in val_a:
                    cutoff_index = i
                    break
            if cutoff_index is not None: df = df.iloc[:cutoff_index]

            debit_col = next((c for c in df.columns if 'дебет' in str(c).lower()), None)
            credit_col = next((c for c in df.columns if 'кредит' in str(c).lower()), None)
            purpose_col = next((c for c in df.columns if 'назначение' in str(c).lower()), None)
            doc_col = next((c for c in df.columns if 'номер док' in str(c).lower() or '№ док' in str(c).lower()), None)

            if debit_col: df.rename(columns={debit_col: 'Оборот Дебет'}, inplace=True)
            if credit_col: df.rename(columns={credit_col: 'Оборот Кредит'}, inplace=True)
            if purpose_col: df.rename(columns={purpose_col: 'Назначение платежа'}, inplace=True)
            if doc_col: df.rename(columns={doc_col: '№ док'}, inplace=True)

            if 'Cчет/ИНН' not in df.columns:
                acc_col = next((c for c in df.columns if 'чет' in str(c).lower() or 'чёт' in str(c).lower()), None)
                name_col = next((c for c in df.columns if 'наименование' in str(c).lower() or 'название' in str(c).lower()), None)
                inn_col = next((c for c in df.columns if 'инн' in str(c).lower()), None)
                if acc_col and name_col:
                    def build_acc_inn(row):
                        acc = str(row[acc_col]).strip() if pd.notna(row[acc_col]) else ""
                        name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                        inn = str(row[inn_col]).strip() if (inn_col and pd.notna(row[inn_col])) else "000000000"
                        if acc == "nan": acc = ""
                        if name == "nan": name = ""
                        if inn == "nan" or not inn: inn = "000000000"
                        if not acc: return None
                        return f"{acc}/{inn}/{name}"
                    df['Cчет/ИНН'] = df.apply(build_acc_inn, axis=1)
                else: raise KeyError("Не удалось найти нужные колонки!")

            df = df[df['Дата'].notna() & df['Cчет/ИНН'].notna() & df['Cчет/ИНН'].str.match(r'^\d+/.+')]
            df['Дата'] = pd.to_datetime(df['Дата'], dayfirst=True, errors='coerce')
            if df.empty: raise ValueError("Нет данных после фильтрации.")
            return df, company_name, period, start_balance

    except Exception as e:
        print(f"Ошибка чтения листа {sheet_name}: {e}")
        return None, None, None, None

# ==========================================
# 2. Логика сводки
# ==========================================
def create_summary(turnover_data, company_name, period, start_balance):
    default_map_expense = {
        "20208": "6010", "16401": "9430", "23106": "6990", "22628": "6530-ИНПС",
        "20216": "4010", "16377": "9610", "13105": "9610", "20214": "4010",
        "23402": "4010", "20816": "6990", "20218": "5830/6810",
    }
    default_map_income = {
        "23404": "4010", "23402": "4010", "20208": "4010", "20214": "4010", "20218": "5830/6810",
    }

    if turnover_data is None: return None

    income_data = []
    expense_data = []

    for _, row in turnover_data.iterrows():
        if pd.notna(row['Cчет/ИНН']) and isinstance(row['Cчет/ИНН'], str):
            acc_raw = str(row['Cчет/ИНН'])
            acc_code = acc_raw[:5] if len(acc_raw) >= 5 else ''
        else: continue

        purpose_text = str(row.get('Назначение платежа', '')).lower()
        debit_val = to_float_safe(row.get('Оборот Дебет', 0))
        credit_val = to_float_safe(row.get('Оборот Кредит', 0))
        mapped_acc = '9430'

        if debit_val > 0:
            found_by_text = False
            if 'абонентская плата' in purpose_text:
                mapped_acc = "6990"
                found_by_text = True
            elif any(word in purpose_text for word in ['природный газ', 'за природный газ']):
                mapped_acc = "6990"
                found_by_text = True
            elif any(word in purpose_text for word in ['согласно договру', 'согласно договору']):
                mapped_acc = "6990"
                found_by_text = True
            elif any(word in purpose_text for word in ['налог на прибыль', 'налог на прибыль(доходы)']):
                mapped_acc = "6410-нал.прибыль"
                found_by_text = True
            elif any(word in purpose_text for word in ['за услуги электронного документооборота', 'электронного документооборота didox.uz', 'за услуги']):
                mapped_acc = "6990"
                found_by_text = True
            elif 'гашение основного долга' in purpose_text:
                mapped_acc = "7810"
                found_by_text = True
            elif 'гашение основных процентов' in purpose_text:
                mapped_acc = "9610"
                found_by_text = True
            elif '09510' in purpose_text:
                mapped_acc = "6990"
                found_by_text = True
            elif 'залоговая сумма' in purpose_text:
                mapped_acc = "6990"
                found_by_text = True

            if not found_by_text:
                if acc_code == "23402":
                    if any(w in purpose_text for w in ['социальный налог']): mapped_acc = "6520-ЕСП"
                    elif any(w in purpose_text for w in ['налог с оборота', 'оплата за 100 налог с оборота']): mapped_acc = "6530-ИНПС"
                    elif any(w in purpose_text for w in ['налог на добавленную стоимость', 'оплата за 1 налог на добавленную стоимость']): mapped_acc = "6410-НДС"
                    elif any(w in purpose_text for w in ['налог на доходы', 'доходы']): mapped_acc = "6410-п/н"
                    else: mapped_acc = "4010"
                elif acc_code == "20208":
                    if 'эл.энергия' in purpose_text: mapped_acc = "6910"
                    else: mapped_acc = "6010"
                elif acc_code == "13105":
                    if any(w in purpose_text for w in ['взыскание просроченного основного долга', 'взыскание просроченных процентов']): mapped_acc = "7810"
                    else: mapped_acc = "9610"
                else: mapped_acc = default_map_expense.get(acc_code, '9430')

        elif credit_val > 0:
            mapped_acc = default_map_income.get(acc_code, '9430')

        entry = {
            'Наименование': acc_raw, 'За что': row.get('Назначение платежа', ''),
            '№ документа': row.get('№ док', ''),
            'Дата': row['Дата'].strftime('%Y-%m-%d %H:%M:%S') if pd.notna(row['Дата']) else '',
            'Оборот по Д-ту': debit_val, 'Оборот по кредиту': credit_val,
        }
        if debit_val > 0: expense_data.append({**entry, mapped_acc: debit_val})
        elif credit_val > 0: income_data.append({**entry, mapped_acc: credit_val})

    expense_df = pd.DataFrame(expense_data)
    income_df = pd.DataFrame(income_data)

    financial_cols_order = [
        '4010', '6990', '5830/6810', '9430', '6010', '6410-п/н', '6410-п/н имущества аренды', '6410-нал.прибыль', '6410-НДС', '6410-ЕНП',
        '6520-ЕСП', '6530-ИНПС', '7810', '9610', '6910', '6410-Налог на воду', '6410-Налог на Землю', '6410-Налог имущества',
        '6610Дивиденды к оплате', '6410-Налог 5% от дивиденда'
    ]

    final_cols = ['Наименование', 'За что', '№ документа', 'Дата'] + financial_cols_order + ['Всего сумма']

    for col in financial_cols_order:
        if col not in expense_df.columns: expense_df[col] = 0.0
        if col not in income_df.columns: income_df[col] = 0.0

    expense_df['Всего сумма'] = None
    income_df['Всего сумма'] = None

    if not expense_df.empty:
        for col in financial_cols_order: expense_df[col] = expense_df[col].apply(lambda x: '' if x == 0.0 else x)
        expense_df = expense_df.reindex(columns=final_cols)

    if not income_df.empty:
        for col in financial_cols_order: income_df[col] = income_df[col].apply(lambda x: '' if x == 0.0 else x)
        income_df = income_df.reindex(columns=final_cols)

    total_expense = sum(d['Оборот по Д-ту'] for d in expense_data)
    total_income = sum(d['Оборот по кредиту'] for d in income_data)
    end_balance = start_balance + total_income - total_expense

    header = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': 'ПРИХОД И РАСХОД ДЕНЕЖНЫХ СРЕДСТВ'}])
    subheader = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': company_name, 'За что': period}])
    cols_row = {c: c for c in final_cols}
    columns_df = pd.DataFrame([cols_row])

    exp_header = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': 'РАСХОД'}])
    if not expense_df.empty:
        exp_total_row = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': 'ИТОГО по расходу', 'Всего сумма': None}])
        expense_block = pd.concat([exp_header, expense_df, exp_total_row], ignore_index=True)
    else: expense_block = exp_header

    inc_header = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': 'ПРИХОД'}])
    if not income_df.empty:
        inc_total_row = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': 'ИТОГО по приходу', 'Всего сумма': None}])
        income_block = pd.concat([inc_header, income_df, inc_total_row], ignore_index=True)
    else: income_block = inc_header

    totals_block = pd.DataFrame([
        {**{c: '' for c in final_cols}, 'Наименование': 'Остаток на начало периода', 'Всего сумма': start_balance},
        {**{c: '' for c in final_cols}, 'Наименование': 'ПРИХОД', 'Всего сумма': total_income},
        {**{c: '' for c in final_cols}, 'Наименование': 'РАСХОД', 'Всего сумма': total_expense},
        {**{c: '' for c in final_cols}, 'Наименование': 'Остаток на конец периода', 'Всего сумма': end_balance}
    ])

    empty_row = pd.DataFrame([{**{c: '' for c in final_cols}, 'Наименование': ''}])
    output_df = pd.concat([header, subheader, columns_df, expense_block, empty_row, income_block, empty_row, totals_block], ignore_index=True)
    output_df = output_df[final_cols].fillna('')

    return output_df

# ==========================================
# 3. Сохранение в Excel с форматированием
# ==========================================
def save_summary_to_excel(summary_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист1'

    for r in dataframe_to_rows(summary_df, index=False, header=False): ws.append(r)

    column_widths = {
        'A': 20, 'B': 30, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12,
        'I': 12, 'J': 12, 'K': 18, 'L': 15, 'M': 12, 'N': 12, 'O': 12, 'P': 12,
        'Q': 12, 'R': 12, 'S': 12, 'T': 15, 'U': 15, 'V': 15, 'W': 15, 'X': 15, 'Y': 18
    }
    for col, width in column_widths.items(): ws.column_dimensions[col].width = width

    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    bold_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:Y1')
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A2'].font = header_font
    ws['A2'].alignment = left_align
    ws['B2'].font = header_font
    ws['B2'].alignment = left_align
    for cell in ws[3]:
        cell.font = header_font
        cell.alignment = center_align

    expense_header_idx = summary_df[summary_df['Наименование'] == 'РАСХОД'].index[0] if not summary_df[summary_df['Наименование'] == 'РАСХОД'].empty else None
    expense_total_idx = summary_df[summary_df['Наименование'] == 'ИТОГО по расходу'].index[0] if not summary_df[summary_df['Наименование'] == 'ИТОГО по расходу'].empty else None
    income_header_idx = summary_df[summary_df['Наименование'] == 'ПРИХОД'].index[0] if not summary_df[summary_df['Наименование'] == 'ПРИХОД'].empty else None
    income_total_idx = summary_df[summary_df['Наименование'] == 'ИТОГО по приходу'].index[0] if not summary_df[summary_df['Наименование'] == 'ИТОГО по приходу'].empty else None

    expense_start_row = (expense_header_idx + 2) if expense_header_idx is not None else None
    expense_end_row = (expense_total_idx) if expense_total_idx is not None else (expense_header_idx + 1 if expense_header_idx is not None else None)
    expense_total_row = (expense_total_idx + 1) if expense_total_idx is not None else None

    income_start_row = (income_header_idx + 2) if income_header_idx is not None else None
    income_end_row = (income_total_idx) if income_total_idx is not None else (income_header_idx + 1 if income_header_idx is not None else None)
    income_total_row = (income_total_idx + 1) if income_total_idx is not None else None

    try:
        totals_start_idx = summary_df[summary_df['Наименование'] == 'Остаток на начало периода'].index[0]
        start_balance_row, total_income_row, total_expense_row, end_balance_row = totals_start_idx + 1, totals_start_idx + 2, totals_start_idx + 3, totals_start_idx + 4
    except IndexError:
        start_balance_row = total_income_row = total_expense_row = end_balance_row = None

    financial_columns_letters = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']

    for row in range(4, ws.max_row + 1):
        is_exp = (expense_start_row <= row < expense_total_row) if (expense_start_row and expense_total_row) else False
        is_inc = (income_start_row <= row < income_total_row) if (income_start_row and income_total_row) else False
        if str(ws[f'A{row}'].value) not in ['РАСХОД', 'ПРИХОД'] and (is_exp or is_inc):
            ws[f'Y{row}'] = f'=SUM(E{row}:X{row})'

    if expense_total_row:
        if expense_start_row and expense_end_row >= expense_start_row:
            for col in financial_columns_letters: ws[f'{col}{expense_total_row}'] = f'=SUM({col}{expense_start_row}:{col}{expense_end_row})'
            ws[f'Y{expense_total_row}'] = f'=SUM(Y{expense_start_row}:Y{expense_end_row})'
        else: ws[f'Y{expense_total_row}'] = 0

    if income_total_row:
        if income_start_row and income_end_row >= income_start_row:
            for col in financial_columns_letters: ws[f'{col}{income_total_row}'] = f'=SUM({col}{income_start_row}:{col}{income_end_row})'
            ws[f'Y{income_total_row}'] = f'=SUM(Y{income_start_row}:Y{income_end_row})'
        else: ws[f'Y{income_total_row}'] = 0

    if total_income_row:
        ws[f'Y{total_income_row}'] = f'=Y{income_total_row}' if income_total_row else 0
        ws[f'Y{total_expense_row}'] = f'=Y{expense_total_row}' if expense_total_row else 0
        ws[f'Y{end_balance_row}'] = f'=Y{start_balance_row} + Y{total_income_row} - Y{total_expense_row}'

    total_rows = list(set([x for x in [expense_total_row, income_total_row, start_balance_row, total_income_row, total_expense_row, end_balance_row] if x is not None]))

    for row_num in total_rows:
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num))
        if row_cells:
            for cell in row_cells[0]:
                cell.border = bold_border
                if cell.column_letter in financial_columns_letters + ['Y']:
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row not in total_rows: cell.border = thin_border
            if cell.column in [1, 2]: cell.alignment = left_align
            elif cell.column in [3, 4]: cell.alignment = center_align
            else:
                cell.alignment = right_align
                if cell.column >= 5 and cell.value != '' and cell.value is not None: cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if str(cell.value) in ['РАСХОД', 'ПРИХОД', 'ИТОГО по расходу', 'ИТОГО по приходу', 'Остаток на начало периода', 'Остаток на конец периода']:
                cell.font = header_font

    wb.save(output_file)

# =========================================================
# ГЛАВНАЯ ФУНКЦИЯ ДЛЯ БОТА
# =========================================================
def process_oborotka_file(file_path, target_month):
    month_name = MONTH_NAMES.get(target_month, "")
    output_files = []

    try:
        with pd.ExcelFile(file_path) as xls:
            sheet_names = xls.sheet_names

        # Создаем словарь для хранения листов с самой поздней датой
        # Формат: {'Название компании': {'end_date': datetime, 'data': turnover_data, ...}}
        best_sheets = {}

        for sheet_name in sheet_names:
            turnover_data, company_name, period, start_balance = read_turnover_sheet(file_path, sheet_name, target_month)

            if turnover_data is not None:
                # Делаем безопасное имя
                safe_name = re.sub(r'[\\/*?:"<>|\n\r\t]', '', str(company_name)).strip()
                if safe_name == "Не указано" or not safe_name:
                    safe_name = f"Неизвестная_фирма_{os.path.basename(file_path).replace('.xlsx', '')}"

                # Извлекаем дату окончания из строки period (которая в формате "DD.MM.YYYY - DD.MM.YYYY")
                try:
                    end_date_str = period.split(' - ')[1] # Берем вторую часть после тире
                    end_date = datetime.strptime(end_date_str, "%d.%m.%Y") # Превращаем в объект даты для сравнения
                except Exception:
                    end_date = datetime.min # Если вдруг дата не прочитается, ставим минимальную

                # Если этой фирмы еще нет в словаре ИЛИ текущая дата окончания больше, чем сохраненная
                if safe_name not in best_sheets or end_date > best_sheets[safe_name]['end_date']:
                    best_sheets[safe_name] = {
                        'end_date': end_date,
                        'data': turnover_data,
                        'company_name': company_name,
                        'period': period,
                        'start_balance': start_balance
                    }

        # Теперь, когда мы отфильтровали самые "свежие" и длинные листы, создаем Excel-сводки
        for safe_name, sheet_info in best_sheets.items():
            output_file = os.path.abspath(f"Сводка_{safe_name}_{target_month}_{month_name}.xlsx")

            summary_df = create_summary(
                sheet_info['data'],
                sheet_info['company_name'],
                sheet_info['period'],
                sheet_info['start_balance']
            )

            if summary_df is not None:
                save_summary_to_excel(summary_df, output_file)
                output_files.append(output_file)

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")

    return output_files